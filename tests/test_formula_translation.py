from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName

from modelwright.extraction import CellRecord, extract_workbook
from modelwright.formulas import build_formula_reference_index, translate_formula_cell
from modelwright.graph import build_dependency_graph
from tests.fixtures.synthetic_model.build_workbook import build_workbook


def synthetic_formula_cells(tmp_path: Path) -> tuple[dict[str, CellRecord], object]:
    workbook = extract_workbook(build_workbook(tmp_path / "synthetic_model.xlsx"))
    graph = build_dependency_graph(workbook)
    return {cell.cell_ref: cell for cell in workbook.cells if cell.formula is not None}, graph


def test_translate_named_range_arithmetic_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Calc!B2"], graph)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "binary"
    assert expression.root.operator == "*"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Inputs!B2"
    assert expression.root.operands[1].kind == "binary"
    assert expression.root.operands[1].operator == "+"
    assert expression.root.operands[1].operands[0].value == 1
    assert expression.root.operands[1].operands[1].reference is not None
    assert expression.root.operands[1].operands[1].reference.normalized == "Inputs!B3"


def test_translate_formula_uses_reference_index(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)
    reference_index = build_formula_reference_index(graph)

    expression = translate_formula_cell(cells["Calc!B2"], graph, reference_index=reference_index)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Inputs!B2"


def test_translate_named_range_backed_by_table_column_as_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "named-range-table-column.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount"])
    sheet.append([10])
    sheet.append([20])
    sheet["C1"] = "=SUM(TableAmounts)"
    sheet.add_table(Table(displayName="InputTable", ref="A1:A3"))
    source.defined_names.add(DefinedName("TableAmounts", attr_text="InputTable[[#Data],[Amount]]"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!C1")

    expression = translate_formula_cell(formula_cell, graph, reference_index=reference_index)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Data!A2:A3"


def test_translate_sheet_relative_arithmetic_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Calc!B3"], graph)

    assert expression.root is not None
    assert expression.root.operator == "*"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Calc!B2"
    assert expression.root.operands[1].reference is not None
    assert expression.root.operands[1].reference.normalized == "Inputs!B4"


def test_translate_round_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Calc!B4"], graph)

    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "ROUND"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.normalized == "Calc!B3"
    assert expression.root.operands[1].value == 2


def test_translate_direct_reference_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Summary!B2"], graph)

    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Calc!B4"


def test_translate_if_formula(tmp_path: Path) -> None:
    cells, graph = synthetic_formula_cells(tmp_path)

    expression = translate_formula_cell(cells["Summary!B3"], graph)

    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "IF"
    assert expression.root.operands[0].kind == "comparison"
    assert expression.root.operands[0].operator == ">"
    assert expression.root.operands[0].operands[0].reference is not None
    assert expression.root.operands[0].operands[0].reference.normalized == "Summary!B2"
    assert expression.root.operands[0].operands[1].value == 50
    assert expression.root.operands[1].value == "ok"
    assert expression.root.operands[2].value == "low"


def test_translate_unsupported_function_reports_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "unsupported_function.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "needle"
    sheet["B1"] = "=XLOOKUP(A1,A:A,B:B)"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_function"
    assert expression.diagnostics[0].severity == "error"
    assert expression.diagnostics[0].raw_value == "XLOOKUP"


def test_translate_unsupported_operator_reports_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "unsupported_operator.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 2
    sheet["B1"] = "=A1%"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_operator"
    assert expression.diagnostics[0].severity == "error"
    assert expression.diagnostics[0].raw_value == "%"


def test_translate_structured_reference_reports_error(tmp_path: Path) -> None:
    workbook_path = tmp_path / "structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["A1"] = "Amount"
    sheet["A2"] = 10
    sheet["B1"] = "=Table1[Amount]"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_structured_reference"
    assert expression.diagnostics[0].severity == "error"
    assert expression.diagnostics[0].raw_value == "Table1[Amount]"


def test_translate_current_row_structured_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "current-row-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, "=InputTable[[#This Row],[Amount]]"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B2"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B2")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Data!A2"


def test_translate_column_structured_reference_as_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "column-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Rate"])
    sheet.append([10, 0.1])
    sheet.append([20, 0.2])
    sheet["D1"] = "=SUM(InputTable[Amount])"
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!D1")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "SUM"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.kind == "range"
    assert expression.root.operands[0].reference.normalized == "Data!A2:A3"


def test_translate_structured_column_span_as_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "column-span-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Selection", "Scenario", "Description"])
    sheet.append(["x", "Current", "same"])
    sheet.append([None, "Future", "changed"])
    sheet["E1"] = '=VLOOKUP("X",InputTable[[Selection]:[Scenario]],2,FALSE)'
    sheet.add_table(Table(displayName="InputTable", ref="A1:C3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!E1")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "VLOOKUP"
    assert expression.root.operands[1].reference is not None
    assert expression.root.operands[1].reference.kind == "range"
    assert expression.root.operands[1].reference.normalized == "Data!A2:B3"


def test_translate_current_row_structured_column_span_as_range(tmp_path: Path) -> None:
    workbook_path = tmp_path / "current-row-column-span-structured-reference.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Start", "End", "Total"])
    sheet.append([10, 2, '=SUM(InputTable[[#This Row],[Start]:[End]])'])
    sheet.add_table(Table(displayName="InputTable", ref="A1:C2"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!C2")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "SUM"
    assert expression.root.operands[0].reference is not None
    assert expression.root.operands[0].reference.kind == "range"
    assert expression.root.operands[0].reference.normalized == "Data!A2:B2"


def test_translate_boolean_literal(tmp_path: Path) -> None:
    workbook_path = tmp_path / "boolean-literal.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "=FALSE"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!A1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "literal"
    assert expression.root.value is False


def test_translate_unary_minus_exponent_and_concat(tmp_path: Path) -> None:
    workbook_path = tmp_path / "operators.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 3
    sheet["A2"] = "x"
    sheet["B1"] = "=-A1"
    sheet["B2"] = "=A1^2"
    sheet["B3"] = '=A2&"y"'
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cells = {cell.cell_ref: cell for cell in workbook.cells if cell.formula is not None}

    unary = translate_formula_cell(formula_cells["Calc!B1"], graph)
    exponent = translate_formula_cell(formula_cells["Calc!B2"], graph)
    concat = translate_formula_cell(formula_cells["Calc!B3"], graph)

    assert unary.translated is True
    assert unary.root is not None
    assert unary.root.kind == "unary"
    assert unary.root.operator == "-"
    assert exponent.translated is True
    assert exponent.root is not None
    assert exponent.root.kind == "binary"
    assert exponent.root.operator == "^"
    assert concat.translated is True
    assert concat.root is not None
    assert concat.root.kind == "binary"
    assert concat.root.operator == "&"


def test_translate_ref_error_emits_sharp_literal(tmp_path: Path) -> None:
    workbook_path = tmp_path / "ref-error.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = "=#REF!"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!A1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is True
    assert expression.root.kind == "literal"
    assert expression.root.value == "#REF!"


def test_translate_static_offset_to_concrete_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "static-offset.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Amount", "Result"])
    sheet.append([10, None])
    sheet.append([20, "=OFFSET(InputTable[[#This Row],[Amount]],-1,0)"])
    sheet.add_table(Table(displayName="InputTable", ref="A1:B3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B3")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Data!A2"


def test_translate_dynamic_offset_shape_reports_sharp_diagnostic(tmp_path: Path) -> None:
    workbook_path = tmp_path / "dynamic-offset.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 10
    sheet["B1"] = "=OFFSET(A1,0,0,2,1)"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_offset_shape"
    assert expression.diagnostics[0].raw_value == "OFFSET"


def test_translate_xlfn_ifna_as_ifna(tmp_path: Path) -> None:
    workbook_path = tmp_path / "xlfn-ifna.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 1
    sheet["B1"] = '=_xlfn.IFNA(A1,"missing")'
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Calc!B1")

    expression = translate_formula_cell(formula_cell, graph)

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "IFNA"


def test_translate_index_match_lookup(tmp_path: Path) -> None:
    workbook_path = tmp_path / "index-match.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Code", "Label", "Result"])
    sheet.append(["A", "Alpha", '=INDEX(InputTable[Label],MATCH(InputTable[[#This Row],[Code]],InputTable[Code],0),0)'])
    sheet.append(["B", "Beta", None])
    sheet.add_table(Table(displayName="InputTable", ref="A1:C3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!C2")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "INDEX"
    assert expression.root.operands[0].kind == "reference"
    assert expression.root.operands[0].reference.kind == "range"
    assert expression.root.operands[0].reference.normalized == "Data!A2:B3" or expression.root.operands[0].reference.normalized == "Data!B2:B3"
    assert expression.root.operands[1].kind == "function_call"
    assert expression.root.operands[1].function_name == "MATCH"


def test_translate_conditional_aggregate_functions(tmp_path: Path) -> None:
    workbook_path = tmp_path / "conditional-aggregates.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Group", "Value"])
    sheet.append(["a", 1])
    sheet.append(["b", 2])
    sheet.append(["a", 3])
    sheet["E1"] = '=AVERAGEIFS(InputTable[Value],InputTable[Group],"a")'
    sheet["E2"] = '=_xlfn.MINIFS(InputTable[Value],InputTable[Group],"b")'
    sheet["E3"] = '=AVERAGEIF(InputTable[Value],">0")'
    sheet.add_table(Table(displayName="InputTable", ref="A1:B4"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)

    for cell_ref, function_name in (("Data!E1", "AVERAGEIFS"), ("Data!E2", "MINIFS"), ("Data!E3", "AVERAGEIF")):
        formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == cell_ref)
        expression = translate_formula_cell(formula_cell, graph, reference_index=reference_index)
        assert expression.translated is True
        assert expression.root.function_name == function_name


def test_translate_numeric_functions(tmp_path: Path) -> None:
    workbook_path = tmp_path / "numeric-functions.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Calc"
    sheet["A1"] = 4
    sheet["B1"] = "=VALUE(\"42\")"
    sheet["B2"] = "=_xlfn.NUMBERVALUE(\"1.5\")"
    sheet["B3"] = "=LN(A1)"
    sheet["B4"] = "=VALUE(A1)"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)

    expected = {"Calc!B1": "VALUE", "Calc!B2": "NUMBERVALUE", "Calc!B3": "LN", "Calc!B4": "VALUE"}
    for cell_ref, function_name in expected.items():
        formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == cell_ref)
        expression = translate_formula_cell(formula_cell, graph, reference_index=reference_index)
        assert expression.translated is True
        assert expression.root.function_name == function_name


def test_translate_repaired_corrupted_structured_reference(tmp_path: Path) -> None:
    workbook_path = tmp_path / "corrupted-structured.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet.append(["Country", "Product"])
    sheet.append(["CA", "Wheat"])
    sheet.append(["US", "Maize"])
    formula = (
        '=INDEX(calc_cropcosts[] calc_cropcosts[[#This Row],[Product]],'
        'MATCH(1,calc_cropcosts[[#This Row],[Country]],0))'
    )
    sheet["D2"] = formula
    sheet.add_table(Table(displayName="calc_cropcosts", ref="A1:C3"))
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!D2")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "function_call"
    assert expression.root.function_name == "INDEX"
    assert expression.root.operands[0].kind == "reference"
    assert expression.root.operands[0].reference.kind == "cell"
    assert expression.root.operands[0].reference.normalized == "Data!B2"


def test_translate_static_indirect_address(tmp_path: Path) -> None:
    workbook_path = tmp_path / "static-indirect.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["A2"] = 42
    sheet["B2"] = "=INDIRECT(ADDRESS(ROW()-1,COLUMN()))"
    sheet["B3"] = "=INDIRECT(ADDRESS(ROW(),COLUMN()-1))"
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    reference_index = build_formula_reference_index(graph)

    b2 = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B2")
    expression = translate_formula_cell(b2, graph, reference_index=reference_index)
    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Data!B1"

    b3 = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B3")
    expression = translate_formula_cell(b3, graph, reference_index=reference_index)
    assert expression.translated is True
    assert expression.root is not None
    assert expression.root.kind == "reference"
    assert expression.root.reference is not None
    assert expression.root.reference.normalized == "Data!A3"


def test_translate_static_indirect_address_with_unsupported_pattern(tmp_path: Path) -> None:
    workbook_path = tmp_path / "nonstatic-indirect.xlsx"
    source = Workbook()
    sheet = source.active
    sheet.title = "Data"
    sheet["A1"] = "X"
    sheet["B2"] = '=INDIRECT("A"&1)'
    source.save(workbook_path)
    workbook = extract_workbook(workbook_path)
    graph = build_dependency_graph(workbook)
    formula_cell = next(cell for cell in workbook.cells if cell.cell_ref == "Data!B2")

    expression = translate_formula_cell(formula_cell, graph, reference_index=build_formula_reference_index(graph))

    assert expression.translated is False
    assert expression.diagnostics[0].code == "unsupported_function"
