from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table


EXPECTED_OUTPUTS = {
    "Calc!B1": 5,
    "Calc!B2": 1,
    "Calc!B3": 6,
    "Calc!B4": 1.5,
    "Calc!B5": 9,
    "Calc!B6": -3,
    "Calc!B7": "xy",
    "Calc!B8": False,
    "Calc!B9": True,
    "Calc!B10": True,
    "Calc!B11": False,
    "Calc!B12": False,
    "Calc!B13": True,
    "Calc!B14": True,
    "Calc!B15": 1.5,
    "Calc!B16": "yes",
    "Calc!B17": True,
    "Calc!B18": True,
    "Calc!B19": True,
    "Calc!B20": 99,
    "Calc!B21": 7,
    "Calc!B22": 1,
    "Calc!B23": 4,
    "Calc!B24": 7 / 3,
    "Calc!B25": "xy3",
    "Calc!B26": 30,
    "TableData!B2": 10,
    "TableData!B3": 20,
}


def build_workbook(path: str | Path) -> Path:
    """Write a small workbook covering currently supported semantics."""

    workbook_path = Path(path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "Inputs"
    calc = workbook.create_sheet("Calc")
    table_data = workbook.create_sheet("TableData")

    inputs["A1"] = "Input"
    inputs["B1"] = "Value"
    inputs["A2"] = "Number"
    inputs["B2"] = 3
    inputs["A3"] = "Text"
    inputs["B3"] = "x"
    inputs["A4"] = "Denominator"
    inputs["B4"] = 2
    inputs["A5"] = "Series 1"
    inputs["B5"] = 1
    inputs["A6"] = "Series 2"
    inputs["B6"] = 2
    inputs["A7"] = "Series 3"
    inputs["B7"] = 4

    calc["A1"] = "Semantics"
    calc["B1"] = "=Inputs!B2+Inputs!B4"
    calc["B2"] = "=Inputs!B2-Inputs!B4"
    calc["B3"] = "=Inputs!B2*Inputs!B4"
    calc["B4"] = "=Inputs!B2/Inputs!B4"
    calc["B5"] = "=Inputs!B2^Inputs!B4"
    calc["B6"] = "=-Inputs!B2"
    calc["B7"] = '=Inputs!B3&"y"'
    calc["B8"] = "=FALSE"
    calc["B9"] = "=Inputs!B2>Inputs!B4"
    calc["B10"] = "=Inputs!B2>=Inputs!B4"
    calc["B11"] = "=Inputs!B2<Inputs!B4"
    calc["B12"] = "=Inputs!B2<=Inputs!B4"
    calc["B13"] = "=Inputs!B2=Inputs!B2"
    calc["B14"] = "=Inputs!B2<>Inputs!B4"
    calc["B15"] = "=ROUND(B4,1)"
    calc["B16"] = '=IF(B9,"yes","no")'
    calc["B17"] = "=TRUE"
    calc["B18"] = "=AND(B9,B10)"
    calc["B19"] = "=OR(B11,B9)"
    calc["B20"] = "=IFERROR(Inputs!B2/0,99)"
    calc["B21"] = "=SUM(Inputs!B5:B7)"
    calc["B22"] = "=MIN(Inputs!B5:B7)"
    calc["B23"] = "=MAX(Inputs!B5:B7)"
    calc["B24"] = "=AVERAGE(Inputs!B5:B7)"
    calc["B25"] = '=CONCATENATE(Inputs!B3,"y",Inputs!B2)'
    calc["B26"] = "=SUM(SemanticsTable[Amount])"

    table_data.append(["Amount", "Result"])
    table_data.append([10, "=SemanticsTable[[#This Row],[Amount]]"])
    table_data.append([20, "=SemanticsTable[[#This Row],[Amount]]"])
    table_data.add_table(Table(displayName="SemanticsTable", ref="A1:B3"))

    workbook.save(workbook_path)
    return workbook_path
